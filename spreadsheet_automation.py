import openpyxl as xl
from openpyxl.chart import BarChart, Reference
#in openpyxl package we use the module called chart and utilizing its 2 classes BarCahrt and Reference


def process_workbook(filename):
    workBook = xl.load_workbook(filename)
    currentSheet = workBook['Sheet1']

    for row in range(2, currentSheet.max_row + 1):
        cell = currentSheet.cell(row,3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = currentSheet.cell(row, 4)
        corrected_price_cell.value = corrected_price


    corrected_values = Reference(currentSheet,
              min_row=2,
              max_row=currentSheet.max_row,
              min_col=4,
              max_col=4)
    currentChart = BarChart()
    currentChart.add_data(corrected_values)
    currentSheet.add_chart(currentChart, 'e2')

    workBook.save(filename)